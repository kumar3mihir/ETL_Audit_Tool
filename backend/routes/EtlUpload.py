# /backend/routes/EtlUpload.py
import os
from numpy import full
import pandas as pd
from flask import Flask, request, jsonify, send_file
from flask import Blueprint
from openai import OpenAI
import json
import uuid
from dotenv import load_dotenv
import zipfile
import xml.etree.ElementTree as ET
from werkzeug.utils import secure_filename  # Import secure_filename
import re
import time
import requests
from tenacity import retry, stop_after_attempt, wait_exponential
import openai  # Ensure you're using the correct API client
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from datetime import datetime

# Define the blueprint
etl_upload_bp = Blueprint("etl_upload", __name__)
load_dotenv()

UPLOAD_FOLDER = "uploads"
EXTRACTED_FOLDER = "extracted_files"
ALLOWED_FILE_EXTENSIONS = ['.py', '.sql', '.txt', '.json', '.xml', '.csv', '.log']  # Define valid ETL script file formats
# OUTPUT_FOLDER = "output"
OUTPUT_FOLDER = "/Users/mihirkumarmallick/Desktop/access_parent/project/output"


os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXTRACTED_FOLDER, exist_ok=True)


def is_valid_etl_file(file_path):
    """Check if the file is an ETL script based on its extension."""
    _, extension = os.path.splitext(file_path)
    return extension.lower() in ALLOWED_FILE_EXTENSIONS


def validate_json(json_file):
    """Try to parse JSON file to check if it is valid."""
    try:
        with open(json_file, 'r') as f:
            json.load(f)  # Try to parse the JSON
        return True
    except (json.JSONDecodeError, IOError):
        return False


def validate_xml(xml_file):
    """Try to parse XML file to check if it is valid."""
    try:
        with open(xml_file, 'r') as f:
            ET.parse(f)  # Try to parse the XML
        return True
    except (ET.ParseError, IOError):
        return False




# script_types = {
#     "py": "Python",
#     "sql": "SQL",
#     "sh": "Shell",
#     "yaml": "YAML",
#     "yml": "YAML",
#     "json": "JSON",
#     "csv": "CSV",
#     "xml": "XML",
#     "java": "Java",
#     "ipynb": "Jupyter Notebook",
#     "bat": "Batch Script",
#     "ps1": "PowerShell Script",
#     "pl": "Perl",
#     "rb": "Ruby",
#     "php": "PHP",
#     "r": "R Script",
#     "scala": "Scala",
#     "go": "Go",
#     "c": "C",
#     "cpp": "C++",
#     "ts": "TypeScript",
#     "js": "JavaScript",
#     "hql": "Hive Query Language",
#     "pig": "Apache Pig Script",
#     "bql": "BigQuery SQL",
#     "spark": "Apache Spark",
#     "pyspark": "PySpark",
#     "airflow": "Apache Airflow DAG",
#     "nifi": "Apache NiFi",
#     "oozie": "Apache Oozie Workflow",
#     "dag": "Directed Acyclic Graph (DAG)",
#     "toml": "TOML Configuration",
#     "ini": "INI Configuration",
#     "properties": "Java Properties File",
#     "cfg": "General Configuration File",
#     "parquet": "Apache Parquet",
#     "avro": "Apache Avro",
#     "orc": "Optimized Row Columnar (ORC)",
#     "xls": "Excel Spreadsheet",
#     "xlsx": "Excel Spreadsheet",
#     "rmd": "R Markdown",
# }

def detect_script_type(file_path):
    """
    Determines the type of ETL script based on the file extension.

    Args:
        file_path (str): Path to the file.

    Returns:
        str: Type of script (Python, SQL, Shell, YAML, JSON, CSV, XML, Java, Jupyter Notebook, or Unknown)
    """
    ext = file_path.lower().split('.')[-1]  # Get file extension in lowercase
    
    script_types = {
        "py": "Python",
        "sql": "SQL",
        "sh": "Shell",
        "yaml": "YAML",
        "yml": "YAML",
        "json": "JSON",
        "csv": "CSV",
        "xml": "XML",
        "java": "Java",
        "ipynb": "Jupyter Notebook",
        "bat": "Batch Script",
        "ps1": "PowerShell Script",
        "pl": "Perl",
        "rb": "Ruby",
        "php": "PHP",
        "r": "R Script",
        "scala": "Scala",
        "go": "Go",
        "c": "C",
        "cpp": "C++",
        "ts": "TypeScript",
        "js": "JavaScript"
    }

    return script_types.get(ext, "Unknown")


def remove_comments(content, file_ext):
    """Removes comments from Python, SQL, JSON, XML, and CSV files."""
    
    if file_ext == "py":
        content = re.sub(r"#.*", "", content)  # Remove Python `#` comments
        content = re.sub(r'""".*?"""', "", content, flags=re.DOTALL)  # Remove Python docstrings
    
    elif file_ext == "sql":
        content = re.sub(r"--.*", "", content)  # Remove SQL `--` comments
        content = re.sub(r"/\*.*?\*/", "", content, flags=re.DOTALL)  # Remove SQL `/* */` comments
    
    elif file_ext == "json":
        content = re.sub(r'//.*', "", content)  # Remove JSON `//` comments
    
    elif file_ext == "xml":
        content = re.sub(r"<!--.*?-->", "", content, flags=re.DOTALL)  # Remove XML `<!-- -->` comments
    
    return content.strip()


def process_file(file_path, file_extension):
    if file_extension == "py":
        with open(file_path, 'r') as f:
            script_content = f.read()
            print(' '.join(script_content.split()[:100]))
        report = analyze_etl_script(script_content, "python", additional_questions="None")
        print(f"Generated report: {report}")
    elif file_extension == "sql":
        with open(file_path, 'r') as f:
            script_content = f.read()
        report = analyze_etl_script(script_content, "sql")
    elif file_extension == "json":
        with open(file_path, 'r') as f:
            json_content = json.load(f)
        report = analyze_etl_script(json.dumps(json_content), "json")
    elif file_extension == "xml":
        with open(file_path, 'r') as f:
            xml_content = f.read()
        report = analyze_etl_script(xml_content, "xml")
    elif file_extension == "csv":
        with open(file_path, 'r') as f:
            csv_content = f.read()
        report = analyze_etl_script(csv_content, "csv")
    else:
        return {"error": f"Unsupported file format: {file_extension}"}

    return report


def read_file_content(file_path):
    """Reads a file and removes comments/documentation to retain only executable code."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        
        file_ext = os.path.splitext(file_path)[1].lower()

        # Comment removal based on file type
        if file_ext == ".py":
            content = re.sub(r"#.*", "", content)  # Remove Python comments
        elif file_ext == ".sql":
            content = re.sub(r"--.*", "", content)  # Remove SQL single-line comments
            content = re.sub(r"/\*[\s\S]*?\*/", "", content)  # Remove SQL multi-line comments
        elif file_ext == ".sh":
            content = re.sub(r"#.*", "", content)  # Remove Shell script comments
        elif file_ext in {".yaml", ".yml"}:
            content = re.sub(r"#.*", "", content)  # Remove YAML comments
        elif file_ext == ".json":
            pass  # JSON does not have inline comments
        elif file_ext == ".xml":
            content = re.sub(r"<!--.*?-->", "", content, flags=re.DOTALL)  # Remove XML comments
        elif file_ext == ".csv":
            pass  # CSV is usually data, no comment removal needed
        elif file_ext == ".java":
            content = re.sub(r"//.*", "", content)  # Remove Java single-line comments
            content = re.sub(r"/\*[\s\S]*?\*/", "", content)  # Remove Java multi-line comments
        elif file_ext == ".ipynb":
            pass  # Jupyter notebooks are JSON-based, no comment removal
        elif file_ext == ".bat":
            content = re.sub(r"::.*", "", content)  # Remove Windows Batch comments
            content = re.sub(r"REM .*", "", content)  # Another way to remove batch comments
        elif file_ext == ".ps1":
            content = re.sub(r"#.*", "", content)  # Remove PowerShell comments
        elif file_ext == ".pl":
            content = re.sub(r"#.*", "", content)  # Remove Perl comments
        elif file_ext == ".rb":
            content = re.sub(r"#.*", "", content)  # Remove Ruby comments
        elif file_ext == ".php":
            content = re.sub(r"//.*", "", content)  # Remove PHP single-line comments
            content = re.sub(r"/\*[\s\S]*?\*/", "", content)  # Remove PHP multi-line comments
        elif file_ext == ".r":
            content = re.sub(r"#.*", "", content)  # Remove R script comments
        elif file_ext == ".scala":
            content = re.sub(r"//.*", "", content)  # Remove Scala single-line comments
            content = re.sub(r"/\*[\s\S]*?\*/", "", content)  # Remove Scala multi-line comments
        elif file_ext == ".go":
            content = re.sub(r"//.*", "", content)  # Remove Go single-line comments
            content = re.sub(r"/\*[\s\S]*?\*/", "", content)  # Remove Go multi-line comments
        elif file_ext == ".c":
            content = re.sub(r"//.*", "", content)  # Remove C single-line comments
            content = re.sub(r"/\*[\s\S]*?\*/", "", content)  # Remove C multi-line comments
        elif file_ext == ".cpp":
            content = re.sub(r"//.*", "", content)  # Remove C++ single-line comments
            content = re.sub(r"/\*[\s\S]*?\*/", "", content)  # Remove C++ multi-line comments
        elif file_ext == ".ts":
            content = re.sub(r"//.*", "", content)  # Remove TypeScript single-line comments
            content = re.sub(r"/\*[\s\S]*?\*/", "", content)  # Remove TypeScript multi-line comments
        elif file_ext == ".js":
            content = re.sub(r"//.*", "", content)  # Remove JavaScript single-line comments
            content = re.sub(r"/\*[\s\S]*?\*/", "", content)  # Remove JavaScript multi-line comments

        print(f"[INFO] Processed file: {file_path} (Cleaned length: {len(content)} chars)")
        return content.strip()

    except Exception as e:
        print(f"[ERROR] Failed to read {file_path}: {e}")
        return ""
    


def traverse_directory(root_dir, allowed_extensions=None):
    """
    Recursively fetch valid ETL script files inside a directory.

    - Only allows files with specific extensions.
    - Skips system, hidden, and non-UTF-8 files.

    Args:
        root_dir (str): Directory to scan.
        allowed_extensions (set): Allowed file extensions.
        
    Returns:
        list: List of valid file paths.
    """
    if allowed_extensions is None:
        allowed_extensions = {
            ".py", ".sql", ".sh", ".txt", ".yaml", ".yml", ".json", ".xml",
            ".csv", ".java", ".ipynb", ".bat", ".ps1", ".pl", ".rb", ".php",
            ".r", ".scala", ".go", ".c", ".cpp", ".ts", ".js"
        }  # Full list of script & config extensions

    all_files = []

    for dirpath, _, filenames in os.walk(root_dir):
        for file in filenames:
            file_path = os.path.join(dirpath, file)

            # Skip hidden/system files
            if file.startswith("."):
                continue  

            # Only allow explicitly listed file extensions
            file_ext = os.path.splitext(file)[1].lower()
            if file_ext not in allowed_extensions:
                continue  

            # Check if file is UTF-8 encoded before adding
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    _ = f.read()  # Validate UTF-8 encoding
                all_files.append(file_path)
            except UnicodeDecodeError:
                continue  

    return all_files


def split_large_script(script, chunk_size=50000):
    """Splits a large script into smaller chunks to avoid API input limits."""
    return [script[i:i+chunk_size] for i in range(0, len(script), chunk_size)]

# new acc zero usage apikey =  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
# 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
# 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
# Configure Nvidia API
client = OpenAI(
    base_url="https://integrate.api.nvidia.com/v1",
    # api_key=os.getenv("apikey_nvidia")
   api_key="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"
)






# prompt = f"""
# You are an *Expert Code & Schema Reviewer* for the Automobile Industry, specializing in *Data Governance, Schema Validation, and Code Audits*.  
# Your task is to analyze the given code content and validate it against the dynamically provided checklist.

# ---

# ### *Validation Logic Based on Checklist Type:*  
# - *If the checklist relates to input data validation*, check for:  
#   - Required attribute checks, length constraints, format checks, reference data validation, range constraints, data type enforcement, uniqueness constraints.  
#   - *Return output strictly based on the given code*, using this format:  
#     {{
#         "Audit Details": [
#             {{ "<column_name_1>": "Validation exists (required attribute, length check, format check, reference data check, range check, DataType check, List of values check, Uniqueness check)", "Audit Result": "Pass", "Audit Reason": "Validation applied correctly" }},
#             {{ "<column_name_2>": "No validation exists", "Audit Result": "Fail", "Audit Reason": "Missing validation" }},
#             {{ "<column_name_3>": "Validation exists (required attribute, reference data check, DataType check)", "Audit Result": "Pass", "Audit Reason": "Partial validation exists" }},
#             {{ "<column_name_4>": "No validation exists", "Audit Result": "Fail", "Audit Reason": "Missing validation" }},
#             {{ "<column_name_5>": "No validation exists", "Audit Result": "Fail", "Audit Reason": "Missing validation" }}
#         ]
#     }}
#     - *If the code does not contain input-related data handling, return:*  
#     {{
#         "Audit Details": [
#             {{"AuditNotneeded": "Code does not have input data entry", "Audit Result": "Pass", "Audit Reason": "No input validation needed" }}
#         ]
#     }}

# ---

# - *If the checklist relates to database constraints, integrity checks, or validation rules*, check for:  
#   - *Primary keys (PRIMARY KEY)*  
#   - *Foreign keys (FOREIGN KEY)*  
#   - *Unique constraints (UNIQUE)*  
#   - *Not null constraints (NOT NULL)*  
#   - *Check constraints (CHECK)*  
#   - *Default values (DEFAULT)*  
#   - *Extract and evaluate constraints strictly from the given code, do not generate example responses.*  
#   - *Return only actual findings from the provided code* in this format:  
#     [
#         {{
#             "Audit Details": [
#                 "1. <column_name> - PRIMARY KEY - Pass",
#                 "2. <column_name> - FOREIGN KEY (Expected: Exists) - Fail",
#                 "3. <column_name> - UNIQUE - Pass",
#                 "4. <column_name> - NOT NULL - Fail"
#             ],
#             "Audit Result": "Fail",
#             "Audit Reason": "Missing required constraints"
#         }}
#     ]

# ---

# - *If the checklist relates to database object names being self-explanatory or user-friendly*, check for:  
#   - Ensure names are clear, descriptive, and follow best practices.  
#   - *Strictly extract names from the given code and do not assume or generate example responses.*  
#   - *Return only actual findings* in this format:  
#     [
#         {{
#             "Audit Details": [
#                 "1. <actual_column_name> - Pass"
#             ],
#             "Audit Result": "Pass",
#             "Audit Reason": "All names user-friendly"
#         }},
#         {{
#             "Audit Details": [
#                 "2.<actual_column_name> - Fail"
#             ],
#             "Audit Result": "Fail",
#             "Audit Reason": "Unclear naming"
#         }}
#     ]

# ---

# - *If the checklist relates to data types in database schema*, check for:  
#   - Ensure the correct data types are used based on column names and intended purpose.  
#   - *Extract actual data types from the code instead of assuming them.*  
#   - *Return only actual findings* in this format:  
#     [
#         {{
#             "Audit Details": [
#                 "1. <column_name> - <actual_data_type> - Pass",
#                 "2. <column_name> - <actual_data_type> - Fail (Expected: <expected_data_type>)"
#             ],
#             "Audit Result": "Fail",
#             "Audit Reason": "Incorrect data type"
#         }}
#     ]

# ---

# - *If the checklist relates to audit data placeholders (who and when data was created, modified, or deleted), check for:*  
#   - Presence of necessary audit columns: CreatedBy, ModifiedBy, CreateDate, ModifiedDate, DeletedBy, DeletedDate.  
#   - *Return only findings based on the actual code* in this format:  
#     [
#         {{
#             "Audit Details": [
#                 "1. CreateDate - Missing - Fail",
#                 "2. CreatedBy - Missing - Fail",
#                 "3. ModifiedDate - Missing - Fail",
#                 "4. ModifiedBy - Missing - Fail",
#                 "5. DeletedDate - Missing - Fail",
#                 "6. DeletedBy - Missing - Fail"
#             ],
#             "Audit Result": "Fail",
#             "Audit Reason": "Audit columns missing"
#         }}
#     ]

# ---

# ### *Checklist for Validation:*  
# {checklist}

# ### *Code Content:*  
# {content}  

# ## *File Path:*
# {file_name}

# ---

# ### *Strict Rules:*  
# 0. *For all the tasks include only input data entry or query or creation column names alone, don't assume python function or methods names has input data entry and ignore the variable names and user-defined function names*
# 1. *Extract and analyze only from the given code. Do not generate example outputs.*  
# 2. *Return only the validation results based on the provided code.*  
# 3. *Do NOT consider file names,function names (def function_name():) and variable names, or general identifiers or API routes as database objects.*  
# 4. *Exclude Flask routes, API handlers, and methods from evaluation.*  
# 5. *Return output ONLY in the required format*—do not add explanations, headings, or extra text.  
# 6. *Ensure the response does NOT include this prompt example*—only return results for the given Code Content.  
# 7. *Extract only column names from input data entry or structured data fields.*  
# 9. *Exclude Flask routes, API handlers, and Python function names from evaluation.*  
# 10. *Return output ONLY in the required format*—do not add explanations, headings, or extra text.  
# 11. *Ignore filenames, executable names (.exe), readme file contents, API routes, and metadata.*  
# 12. *Ensure the response does NOT include this prompt example*—only return results for the given Code Content.  
# 13. *Keep audit reasoning concise (≤ 8 words).*  
# 14. *Avoid redundant JSON keywords like 'json' or '```json' in the output.* 
# """



# user_input = ["Question":]
# API Retry Logic: Retries up to 3 times with exponential backoff
@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=10))
def call_genai_api(prompt):
    print("Calling GenAI API...")

    completion = client.chat.completions.create(
        model="meta/llama-3.3-70b-instruct",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.3,
        top_p=0.5,
        max_tokens=4096,
        stop=["Let's analyze", "Starting with"],
        stream=True
    )
    
    return completion  # Return API response



# def analyze_etl_script(script_content, additional_questions=None):
#     """Sends the cleaned script to GenAI for audit analysis and extracts structured JSON output."""
    
#     print("Entering analyze_etl_script() function...")
#     print(f"Script Content (first 100 chars): {script_content[:100]}")

#     prompt = f"""
#     You are an ETL audit expert. Analyze the following ETL script for compliance and based on the checklist user given by prompt:
    
#     {additional_questions}
    
#     ETL Checklist1:"If Question Asked in the additional question is that that check this whole ETL pipeline has multi staging layers or not between source to target data warehouse"
    
#     give respone in this format 
     
#     Your annswer will be in the form of 
#     Audit Report for Data Pipeline Design (CSV Format)
# # Audit Start Date & Time	Audit End Date & Time	File Type	File Name Full Path	Audit Status	Audit Result	Audit Details	Audit Result Reason
# # 22-03-2025 16:00	22-03-2025 16:05	SQL / Python	C:\ETL\landing_zone.sql	Completed	Pass	1. Landing - Raw data stored without transformation	Data retained successfully
# # 22-03-2025 16:00	22-03-2025 16:05	SQL / Python	C:\ETL\staging_layer.py	Completed	Pass	2. Staging - Format standardization & de-duplication applied	Cleaning process verified
# # 22-03-2025 16:00	22-03-2025 16:05	SQL / Python	C:\ETL\integration_layer.sql	Completed	Pass 3. Integration - Matching rules implemented correctly	Entity mapping successful
# # 22-03-2025 16:00	22-03-2025 16:05	SQL / Python	C:\ETL\transformation.py	Completed	Fail	4. Transformation - Summarization logic incorrect	Incorrect data aggregation
# # 22-03-2025 16:00	22-03-2025 16:05	SQL / Python	C:\ETL\warehouse_load.sql	Completed	Fail	5. Final Layer - Target schema mismatch	Schema validation failed
#     I
    
    
    
    
#    ETL Checklist2:"If Question Asked in the additional question is that that check this whole ETL pipeline has proper coding standard like Auditability and reconcilability , restatrtability and exception handling etc "
    
# # •  ETL Process Compliance Audit Report (for ETL Audit)
# # Example ETL Audit Report (CSV format)
# # Audit Start Date & Time	Audit End Date & Time	File Type	File Name Full Path	Audit Status	Audit Result	Audit Details	Audit Result Reason
# # 22-03-2025 16:00	22-03-2025 16:05	Python	C:\ETL\transform.py	Completed	Pass	1. Auditability - Start/End timestamps logged	Log entries found
# # 22-03-2025 16:00	22-03-2025 16:05	Python	C:\ETL\transform.py	Completed	Pass	2. Reconcilability - Row counts match source & target	Data validation passed
# # 22-03-2025 16:00	22-03-2025 16:05	Python	C:\ETL\transform.py	Completed	Fail	3. Restartability - No checkpoint mechanism found	Restart logic missing
# # 22-03-2025 16:00	22-03-2025 16:05	Python	C:\ETL\transform.py	Completed	Fail	4. Exception Handling - Missing try-except block for critical operation	Error handling missing
#     {File_Name}:
#     {script_content}
    
#     """

#     # If additional questions are provided, add them
#     if additional_questions:
#         prompt += f"\n\nAdditionally, answer the following questions:\n{additional_questions}"

#     # Enforce structured JSON format
#     prompt += """
    
#     **Provide structured JSON output at the end in this format and in evidence section add all that have given in the response i.e, all the technical terms and area of improvement:**
#     'so i am using the below to extract output so be serious about generating this output'
#     structured_match = re.search(r"```structured-results\n({.*?})\n```", audit_report, re.DOTALL)

#     ```structured-results
#     {
     
#     }
    
#          } 
#     }
#     ```
#     **DO NOT** include any extra text outside of the JSON output.
#     """

#     print("Generated Prompt:")
#     print(prompt[:300])  # Print only the first 300 characters for debugging

#     try:
#         completion = call_genai_api(prompt)  # Call API with retry logic
#         print("Received response from GenAI model.")

#         audit_report = ""
#         for chunk in completion:
#             # if chunk["choices"][0]["delta"].get("content") is not None:
#             #     print(chunk["choices"][0]["delta"]["content"], end="")
#             #     audit_report += chunk["choices"][0]["delta"]["content"]
#             if chunk.choices[0].delta.content is not None:
#               print(chunk.choices[0].delta.content, end="")
#               audit_report += chunk.choices[0].delta.content
                

#         # Extract structured JSON using regex
#         structured_match = re.search(r"```structured-results\n({.*?})\n```", audit_report, re.DOTALL)

#         if structured_match:
#             structured_json_str = structured_match.group(1)
#             try:
#                 structured_results = json.loads(structured_json_str)
#                 return structured_results  # Return structured JSON
#             except json.JSONDecodeError:
#                 print("[ERROR] Failed to parse structured JSON.")
#                 return {"error": "Failed to parse structured JSON."}

#         print("[WARNING] No structured JSON found! Returning full text instead.")
#         return {"error": "AI did not return structured JSON", "raw_text": audit_report}

#     except Exception as e:
#         print("ERROR OCCURRED in analyze_etl_script:", str(e))
#         return {"error": str(e) + "\n\nPlease check the input and try again."}



import re
import json

# def analyze_etl_script(script_content, file_name, additional_questions=None):
#     """Analyzes an ETL script for compliance and extracts structured audit results in JSON format."""
    
#     print("Entering analyze_etl_script() function...")
#     print(f"Analyzing File: {file_name}")
    
#     prompt = f"""
#     You are an ETL audit expert. Analyze the given ETL script and assess its compliance based on best practices.
    
#     **Checklist for ETL Compliance:**
#     1️⃣ Multi-Staging Layers: Check if the pipeline follows a structured data flow.
#     2️⃣ Coding Standards: Ensure auditability, reconcilability, restartability, and exception handling.

#     **Audit Report Format:**
#     - Audit Start Date & Time: When analysis begins
#     - Audit End Date & Time: When analysis ends
#     - File Type: SQL / Python
#     - File Name Full Path: {file_name}
#     - Audit Status: Completed / Failed
#     - Audit Result: Pass / Partial / Fail
#     - Audit Details: What was checked
#     - Audit Result Reason: Why it passed or failed

#     **Example Output (JSON Format):**
#     ```structured-results
#     {{
#         "file_name": "{file_name}",
#         "audit_details": [
#             {{
#                 "audit_detail": "Landing Zone - Raw data stored without transformation",
#                 "audit_result": "Pass",
#                 "audit_reason": "Data retained successfully"
#             }},
#             {{
#                 "audit_detail": "Transformation - Summarization logic incorrect",
#                 "audit_result": "Fail",
#                 "audit_reason": "Incorrect data aggregation"
#             }}
#         ]
#     }}
#     ```

#     **DO NOT** add extra text before or after the JSON output.

#     **ETL Script Content:**  
#     {script_content}
#     """

#     # If additional questions are provided, add them
#     if additional_questions:
#         prompt += f"\n\nAdditionally, answer the following questions:\n{additional_questions}"

#     print("Generated Prompt:")
#     print(prompt[:300])  # Print only the first 300 characters for debugging

#     try:
#         completion = call_genai_api(prompt)  # Call AI model (Hugging Face, OpenRouter, etc.)
#         print("Received response from GenAI model.")

#         audit_report = ""
#         for chunk in completion:
#             if chunk.choices[0].delta.content is not None:
#                 print(chunk.choices[0].delta.content, end="")
#                 audit_report += chunk.choices[0].delta.content

#         # Extract structured JSON using regex
#         structured_match = re.search(r"```structured-results\n({.*?})\n```", audit_report, re.DOTALL)

#         if structured_match:
#             structured_json_str = structured_match.group(1)
#             try:
#                 structured_results = json.loads(structured_json_str)
#                 return structured_results  # Return properly formatted JSON
#             except json.JSONDecodeError:
#                 print("[ERROR] Failed to parse structured JSON.")
#                 return {"error": "Failed to parse structured JSON."}

#         print("[WARNING] No structured JSON found! Returning full text instead.")
#         return {"error": "AI did not return structured JSON", "raw_text": audit_report}

#     except Exception as e:
#         print("ERROR OCCURRED in analyze_etl_script:", str(e))
#         return {"error": str(e) + "\n\nPlease check the input and try again."}

def format_audit_results(reports):
    """Formats the audit results to match the required CSV format."""
    
    formatted_report = []

    for report in reports:
        if "file_name" in report and "audit_results" in report:
            file_name = report["file_name"]
            for audit_entry in report["audit_results"]:
                formatted_report.append({
                    "File Name": file_name,
                    "Audit Detail": audit_entry.get("Audit Detail", ""),
                    "Audit Result Reason": audit_entry.get("Audit Result Reason", ""),
                    "Audit Status": audit_entry.get("Audit Status", ""),
                })

    return formatted_report


import re
import json

def analyze_etl_script(file_name, script_content, additional_questions=None):
    """Sends the cleaned script to GenAI for audit analysis and extracts structured JSON output."""

    print(f"📌 [INFO] Analyzing file: {file_name}")

    # Prompt preparation
    prompt = f"""
    You are an ETL audit expert. Analyze the following ETL script for compliance based on the checklist:

    {additional_questions}

    **File Name:** {file_name}

    **Script Content:**
    ```
    {script_content}
    ```

    **Provide structured JSON output in this format:**
    ```structured-results
    {{
        "file_name": "{file_name}",
        "audit_results": [
            {{
                "Audit Detail": "Auditability - Start/End timestamps logged",
                "Audit Result Reason": "Log entries found",
                "Audit Status": "Pass"
            }},
            {{
                "Audit Detail": "Exception Handling - Missing try-except block",
                "Audit Result Reason": "Error handling missing",
                "Audit Status": "Fail"
            }}
        ]
    }}
    ```
    **DO NOT** include any extra text outside of the JSON output.
    """

    try:
        print("🚀 [INFO] Sending request to GenAI API...")
        completion = call_genai_api(prompt)  # Your function to call the API

        print("📡 [INFO] Receiving streamed response from GenAI:")
        audit_report = ""

        # Streaming handling
        for chunk in completion:
            chunk_text = chunk.choices[0].delta.content
            if chunk_text:
                print(f"🔹{chunk_text}")  # Debugging each streamed token
                audit_report += chunk_text

        print("✅ [INFO] Received full response from GenAI.")

        # Extract structured JSON
        structured_match = re.search(r"```structured-results\n({.*?})\n```", audit_report, re.DOTALL)
        
        if structured_match:
            structured_json_str = structured_match.group(1)
            print(f"📑 [DEBUG] Extracted JSON String:\n{structured_json_str}")

            try:
                structured_results = json.loads(structured_json_str)
                print("✅ [INFO] Successfully parsed structured JSON.")
                return structured_results
            except json.JSONDecodeError as e:
                print(f"❌ [ERROR] Failed to parse structured JSON: {e}")
                return {"error": "Failed to parse structured JSON.", "raw_text": audit_report}

        print("⚠️ [WARNING] AI did not return structured JSON.")
        return {"error": "AI did not return structured JSON", "raw_text": audit_report}

    except Exception as e:
        print(f"❌ [ERROR] API call failed: {str(e)}")
        return {"error": str(e)}




@etl_upload_bp.route("/upload", methods=["POST"])
def upload_file():
    """Handles file upload and processes ZIP contents if necessary."""
    
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file provided"}), 400

    file_ext = file.filename.split(".")[-1].lower()
    extracted_files = {}

    if file_ext == "zip":
        with zipfile.ZipFile(file, "r") as zip_ref:
            for zip_info in zip_ref.infolist():
                if not zip_info.is_dir():  # Ignore directories
                    with zip_ref.open(zip_info.filename) as extracted_file:
                        extracted_files[zip_info.filename] = extracted_file.read().decode(errors="ignore")
        
        if not extracted_files:
            return jsonify({"error": "ZIP extracted but contains no valid files"}), 400
    else:
        extracted_files[file.filename] = file.read().decode(errors="ignore")

    return jsonify({
        "message": "Files processed successfully.",
        "latest_files": extracted_files  # Dictionary of {filename: content}
    }), 200





#Modified code 4 -- getting strucutred json output
audit_results_cache = {}

@etl_upload_bp.route("/audit", methods=["POST"])
def audit_etl():
    """Handles ETL audit for uploaded files and returns structured JSON output."""
    
    data = request.json
    print("🔥 Incoming JSON Data:", data)  # DEBUG PRINT

    latest_files = data.get("latest_files", {})  # Expecting a dictionary {filename: content}
    additional_questions = data.get("additional_questions", None)

    if not latest_files:
        print("🚨 [ERROR] No files received!")
        return jsonify({"error": "No file provided"}), 400

    print(f"✅ Received {len(latest_files)} file(s): {list(latest_files.keys())}")  # DEBUG PRINT

    full_report = []
    
    for file_name, script_content in latest_files.items():
        if not script_content.strip():
            print(f"[WARNING] Skipping empty file: {file_name}")
            continue

        print(f"Processing {file_name} (Characters: {len(script_content)})")  # DEBUG PRINT
        script_chunks = split_large_script(script_content)

        for chunk in script_chunks:
            audit_result = analyze_etl_script(file_name, chunk, additional_questions)
            if "error" in audit_result:
                print("[ERROR] API call failed:", audit_result["error"])
                return jsonify({"error": "Failed to analyze script"}), 500

            full_report.append(audit_result)

    structured_report = format_audit_results(full_report)
    print("✅ Final Report:", structured_report)  # DEBUG PRINT

    return jsonify({"structured_audit_report": structured_report}), 200


# audit_results_cache = {}
# @etl_upload_bp.route("/audit", methods=["POST"])
# def audit_etl():
#     """Handles ETL audit for uploaded files and returns structured JSON output."""
#     data = request.json
#     latest_files = data.get("latest_files", [])
#     test_mode = data.get("test_mode", False)
#     additional_questions = data.get("additional_questions", None)  # Capture additional questions

#     if not latest_files:
#         print("[ERROR] No files provided for audit.")
#         return jsonify({"error": "No files provided for audit"}), 400

#     combined_script_content = ""
    
#     for file_path in latest_files:
#         if not os.path.exists(file_path):
#             print(f"[WARNING] Skipping missing file: {file_path}")
#             continue

#         script_content = read_file_content(file_path)

#         if script_content:
#             combined_script_content += f"\n### File: {file_path} ###\n{script_content}\n"

#     if not combined_script_content.strip():
#         print("[ERROR] No valid files to process.")
#         return jsonify({"error": "No valid files to process"}), 400

#     # Split script if too large
#     script_chunks = split_large_script(combined_script_content)

#     # If test mode is enabled, return a small chunk to check API response
#     if test_mode:
#         test_prompt = f"Test this script for ETL compliance:\n\n{script_chunks[0]}"
#         return jsonify({"test_prompt": test_prompt}), 200

#     full_report = []
    
    
#     # // we need to sumerise it which code is not implemented till now we can do 
#     for chunk in script_chunks:
#         audit_result = analyze_etl_script(chunk, additional_questions)
        
#         if "error" in audit_result:
#             print("[ERROR] API call failed:", audit_result["error"])
#             return jsonify({"error": "Failed to analyze script"}), 500
        
        
#         full_report.append(audit_result)  # Append structured JSON

#     # Merge multiple JSON results
#     final_report = {
#         "Auditability": {"result": [], "evidence": []},
#         "Reconcilability": {"result": [], "evidence": []},
#         "Restartability": {"result": [], "evidence": []},
#         "Exception Handling": {"result": [], "evidence": []},
#         "Script Contains Only Comments/Readme": {"result": [], "evidence": []},
#         "Follows Best Practices": {"result": [], "evidence": []},
#         "Additional Questions": {}  # Stores questions as key-value pairs (No result/evidence structure)
#     }

#     # for report in full_report:
#     #     for key in report:
#     #         if key == "Additional Questions":
#     #             final_report[key].update(report[key])  # Directly store key-value answers
#     #         elif key in final_report:
#     #             final_report[key]["result"].append(report[key].get("result", "N/A"))
#     #             final_report[key]["evidence"].append(report[key].get("evidence", "N/A"))
#     #         else:
#     #             print(f"⚠️ Warning: Unexpected key '{key}' in report.")
    
#     for report in full_report:
#       for key in report:
#         if key == "Additional Questions":
#             # Ensure each question follows structured result/evidence format
#             for question, details in report[key].items():
#                 if question not in final_report[key]:  # First time seeing this question
#                     final_report[key][question] = {"result": [], "evidence": []}
                
#                 final_report[key][question]["result"].append(details.get("result", "N/A"))
#                 final_report[key][question]["evidence"].append(details.get("evidence", "N/A"))

#         elif key in final_report:
#             final_report[key]["result"].append(report[key].get("result", "N/A"))
#             final_report[key]["evidence"].append(report[key].get("evidence", "N/A"))
#         else:
#             print(f"⚠️ Warning: Unexpected key '{key}' in report.")

#     print("✅ Final Report:", final_report)
#     audit_results_cache["final"] = final_report  

#     return jsonify({"structured_audit_report": final_report}), 200


# #Modified code 3 -- getting strucutred json output
# @etl_upload_bp.route("/audit", methods=["POST"])
# def audit_etl():
#     """Handles ETL audit for uploaded files and returns structured JSON output."""
#     data = request.json
#     latest_files = data.get("latest_files", [])
#     test_mode = data.get("test_mode", False)
#     additional_questions = data.get("additional_questions", None)  # Capture additional questions

#     if not latest_files:
#         print("[ERROR] No files provided for audit.")
#         return jsonify({"error": "No files provided for audit"}), 400

#     combined_script_content = ""
    
#     for file_path in latest_files:
#         if not os.path.exists(file_path):
#             print(f"[WARNING] Skipping missing file: {file_path}")
#             continue

#         script_content = read_file_content(file_path)

#         if script_content:
#             combined_script_content += f"\n### File: {file_path} ###\n{script_content}\n"

#     if not combined_script_content.strip():
#         print("[ERROR] No valid files to process.")
#         return jsonify({"error": "No valid files to process"}), 400

#     # Split script if too large
#     script_chunks = split_large_script(combined_script_content)

#     # If test mode is enabled, return a small chunk to check API response
#     if test_mode:
#         test_prompt = f"Test this script for ETL compliance:\n\n{script_chunks[0]}"
#         return jsonify({"test_prompt": test_prompt}), 200

#     full_report = []
    
#     for chunk in script_chunks:
#         audit_result = analyze_etl_script(chunk, additional_questions)
        
#         if "error" in audit_result:
#             print("[ERROR] API call failed:", audit_result["error"])
#             return jsonify({"error": "Failed to analyze script"}), 500
        
#         full_report.append(audit_result)  # Append structured JSON

#     # audit_results_cache["latest"] = full_report
    
#     # Merge multiple JSON results
#     final_report = {
#         "Auditability": {"result": [], "evidence": []},
#         "Reconcilability": {"result": [], "evidence": []},
#         "Restartability": {"result": [], "evidence": []},
#         "Exception Handling": {"result": [], "evidence": []},
#         "Script Contains Only Comments/Readme": {"result": [], "evidence": []},
#         "Follows Best Practices": {"result": [], "evidence": []},
#         "Additional Questions": {}  # Placeholder for additional questions
#     }

#     for report in full_report:
#         for key in final_report.keys():
#             if key in report:
#                 final_report[key]["result"].append(report[key]["result"])
#                 final_report[key]["evidence"].append(report[key]["evidence"])

#         # Merge additional questions
#         if "Additional Questions" in report:
#             final_report["Additional Questions"].update(report["Additional Questions"])
#     print("final_report", final_report)
#     audit_results_cache["final"] = final_report  
#     return jsonify({"structured_audit_report": final_report}), 200




import os
import csv
from io import BytesIO
from fpdf import FPDF
from flask import send_file, jsonify

# Directory to store generated reports
REPORTS_DIR = "reports"
os.makedirs(REPORTS_DIR, exist_ok=True)

def generate_csv_report(audit_data):
    """Generates a CSV file from audit results and returns the file path."""
    csv_filename = os.path.join(REPORTS_DIR, "audit_report.csv")

    with open(csv_filename, mode="w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(["Category", "Result", "Detailed Analysis", "Technical Improvements"])

        for category, details in audit_data.items():
            if category == "Additional Questions":
                for question, q_details in details.items():
                    evidence = q_details.get("evidence", {})
                    writer.writerow([
                        question,
                        ", ".join(q_details.get("result", ["N/A"])),  # Prevent empty result
                        evidence.get("Detailed Analysis", "N/A"),
                        evidence.get("Technical Improvements", "N/A")
                    ])
            else:
                evidence = details.get("evidence", {})
                writer.writerow([
                    category,
                    ", ".join(details.get("result", ["N/A"])),
                    evidence.get("Detailed Analysis", "N/A"),
                    evidence.get("Technical Improvements", "N/A")
                ])
    return csv_filename

def generate_pdf_report(audit_data):
    """Generates a PDF file from audit results and returns the file path."""
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(200, 10, "ETL Audit Report", ln=True, align="C")
    pdf.ln(10)

    for category, details in audit_data.items():
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, f"{category}", ln=True, align="L")
        pdf.set_font("Arial", "", 10)

        if category == "Additional Questions":
            for question, q_details in details.items():
                evidence = q_details.get("evidence", {})
                pdf.set_font("Arial", "B", 10)
                pdf.multi_cell(0, 6, f"🔹 {question}: {', '.join(q_details.get('result', ['N/A']))}")
                pdf.set_font("Arial", "", 10)
                pdf.multi_cell(0, 6, f"📌 Detailed Analysis: {evidence.get('Detailed Analysis', 'N/A')}")
                pdf.multi_cell(0, 6, f"🛠️ Technical Improvements: {evidence.get('Technical Improvements', 'N/A')}")
                pdf.ln(5)
        else:
            evidence = details.get("evidence", {})
            pdf.multi_cell(0, 6, f"🔹 Result: {', '.join(details.get('result', ['N/A']))}")
            pdf.multi_cell(0, 6, f"📌 Detailed Analysis: {evidence.get('Detailed Analysis', 'N/A')}")
            pdf.multi_cell(0, 6, f"🛠️ Technical Improvements: {evidence.get('Technical Improvements', 'N/A')}")
            pdf.ln(5)

    pdf_filename = os.path.join(REPORTS_DIR, "audit_report.pdf")
    pdf.output(pdf_filename)
    return pdf_filename

# ✅ CSV Download API
@etl_upload_bp.route("/download/csv", methods=["GET"])
def download_csv():
    """Endpoint to download audit results as CSV."""
    if "final" not in audit_results_cache:
        return jsonify({"error": "No audit report found"}), 404

    csv_path = generate_csv_report(audit_results_cache["final"])
    return send_file(csv_path, as_attachment=True, download_name="audit_report.csv")

# ✅ PDF Download API
@etl_upload_bp.route("/download/pdf", methods=["GET"])
def download_pdf():
    """Endpoint to download audit results as PDF."""
    if "final" not in audit_results_cache:
        return jsonify({"error": "No audit report found"}), 404

    pdf_path = generate_pdf_report(audit_results_cache["final"])
    return send_file(pdf_path, as_attachment=True, download_name="audit_report.pdf")

